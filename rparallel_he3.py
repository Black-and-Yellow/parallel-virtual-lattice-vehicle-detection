import cv2
import numpy as np
import psutil
import time
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
import cProfile
import pstats
from concurrent.futures import ThreadPoolExecutor
import json

"""
rparallel_he3.py

Usage notes:
- Set `video` in `user_input_data.json` to either a video file path OR a directory containing
    an image sequence (e.g., a UA-DETRAC sequence folder like `DETRAC-Images/DETRAC-Images/MVI_20011`).
- The script will read consecutive frames (pairwise) and process them the same as before.
    If you want to use UA-DETRAC annotations (XML), we can add parsing and filtering in a follow-up.
"""


# Load configuration from JSON file
with open("user_input_data.json", "r") as file:
    config = json.load(file)

# Extract parameters
video_path = config["video"]
color_channel = config["color_channel"]
rows, cols = config["grids"]["rows"], config["grids"]["cols"]


roi1_x, roi1_y, roi1_width, roi1_height = 545, 159, 284, 140
roi2_x, roi2_y, roi2_width, roi2_height = 238, 161, 284, 140

num_rows, num_cols = rows,cols
grid_width1 = roi1_width // num_cols
grid_height1 = roi1_height // num_rows
grid_width2 = roi2_width // num_cols
grid_height2 = roi2_height // num_rows

density_values_lane1 = []
density_values_lane2 = []

frame_count = 0
start_time = time.time()

result_matrix1 = np.zeros((num_rows, num_cols), dtype=int)
result_matrix2 = np.zeros((num_rows, num_cols), dtype=int)

# Support either a video file or a directory of images (UA-DETRAC sequences)
def load_image_sequence(path):
    exts = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff')
    files = [f for f in os.listdir(path) if f.lower().endswith(exts)]
    files.sort()
    return [os.path.join(path, f) for f in files]

is_image_sequence = os.path.isdir(video_path)
image_files = []
cap = None
if is_image_sequence:
    image_files = load_image_sequence(video_path)
    if len(image_files) < 2:
        raise ValueError(f"Image sequence at {video_path} must contain at least 2 images")
    first_img = cv2.imread(image_files[0])
    h, w = first_img.shape[:2]
else:
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise ValueError(f"Cannot open video file: {video_path}")
    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

fourcc = cv2.VideoWriter_fourcc(*'mp4v')
out = cv2.VideoWriter('test_output1.mp4', fourcc, 20.0, (w, h))

# Create a generator to yield successive frame pairs regardless of source
def frame_pair_generator():
    if is_image_sequence:
        for i in range(len(image_files) - 1):
            f1 = cv2.imread(image_files[i])
            f2 = cv2.imread(image_files[i + 1])
            if f1 is None or f2 is None:
                continue
            yield f1, f2
    else:
        ret, f1 = cap.read()
        if not ret:
            return
        ret, f2 = cap.read()
        while ret and f1 is not None and f2 is not None:
            yield f1, f2
            f1 = f2
            ret, f2 = cap.read()

excel_file_path = 'result_matrix1.xlsx'

def append_to_excel(result_matrix):
    try:
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"Error loading workbook: {e}. Creating a new one.")
        workbook = Workbook()
        sheet = workbook.active

    result_df = pd.DataFrame(result_matrix)

    next_row = sheet.max_row + 2 if sheet.max_row > 1 else 1 

    for row_index, row in enumerate(result_df.values):
        for col_index, value in enumerate(row):
            sheet.cell(row=next_row + row_index, column=col_index + 1, value=value)

    workbook.save(excel_file_path)

def process_channel(channel):
    blur = cv2.GaussianBlur(channel, (5, 5), 0)
    _, thresh = cv2.threshold(blur, 50, 255, cv2.THRESH_BINARY)
    dilated = cv2.dilate(thresh, None, iterations=3)
    contours, _ = cv2.findContours(dilated, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    return contours

def apply_histogram_equalization_single(frame, roi_x, roi_y, roi_width, roi_height):
    roi = frame[roi_y:roi_y + roi_height, roi_x:roi_x + roi_width]
    roi_equalized = cv2.equalizeHist(cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY))
    frame[roi_y:roi_y + roi_height, roi_x:roi_x + roi_width] = cv2.cvtColor(roi_equalized, cv2.COLOR_GRAY2BGR)
    return frame

def apply_histogram_equalization_parallel(frame):
    with ThreadPoolExecutor() as executor:
        futures = [
            executor.submit(apply_histogram_equalization_single, frame, roi1_x, roi1_y, roi1_width, roi1_height),
            executor.submit(apply_histogram_equalization_single, frame, roi2_x, roi2_y, roi2_width, roi2_height)
        ]
        for future in futures:
            frame = future.result()
    return frame

def process_hsv(frame1, frame2, channels):
    frame1 = apply_histogram_equalization_parallel(frame1)
    frame2 = apply_histogram_equalization_parallel(frame2)
    
    diff = cv2.absdiff(frame1, frame2)
    hsv = cv2.cvtColor(diff, cv2.COLOR_BGR2HSV)
    channels_data = [cv2.split(hsv)[i] for i in channels]
    return channels_data

def process_grayscale(frame1, frame2):
    frame1 = apply_histogram_equalization_parallel(frame1)
    frame2 = apply_histogram_equalization_parallel(frame2)
    
    diff = cv2.absdiff(frame1, frame2)
    gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
    return [gray]

def process_grid_cell(args):
    roi_x, roi_y, grid_width, grid_height, frame, channels_data = args
    result = 0
    detection_flags = []
    for channel in channels_data:
        grid_channel = channel[roi_y:roi_y + grid_height, roi_x:roi_x + grid_width]
        contours = process_channel(grid_channel)
        detection_flags.append(any(cv2.contourArea(contour) >= 100 for contour in contours))
    if all(detection_flags):
        result = 1
    return result

def process_grid(roi_x, roi_y, grid_width, grid_height, result_matrix, channels_data, frame):
    with ThreadPoolExecutor() as executor:
        args = [(roi_x + col * grid_width, roi_y + row * grid_height, grid_width, grid_height, frame, channels_data)
                for row in range(num_rows) for col in range(num_cols)]
        results = executor.map(process_grid_cell, args)
        for index, result in enumerate(results):
            row = index // num_cols
            col = index % num_cols
            result_matrix[row, col] = result

user_choice = color_channel

choices = {
    'H': [0],
    'S': [1],
    'V': [2],
    'H+S': [0, 1],
    'H+V': [0, 2],
    'S+V': [1, 2],
    'H+S+V': [0, 1, 2],
    'gray': 'gray'
}

channels = choices[user_choice]

def main():
    global frame_count
    for f1, f2 in frame_pair_generator():
        frame_count += 1

        if f1.shape[:2] != f2.shape[:2]:
            # Skip mismatched frames
            continue

        if channels == 'gray':
            channels_data = process_grayscale(f1, f2)
        else:
            channels_data = process_hsv(f1, f2, channels)

        result_matrix1.fill(0)
        result_matrix2.fill(0)

        process_grid(roi1_x, roi1_y, grid_width1, grid_height1, result_matrix1, channels_data, f1)
        process_grid(roi2_x, roi2_y, grid_width2, grid_height2, result_matrix2, channels_data, f1)

        # append_to_excel(result_matrix1)

        for row in range(num_rows):
            for col in range(num_cols):
                if result_matrix1[row, col] == 0:
                    grid_x = roi1_x + col * grid_width1
                    grid_y = roi1_y + row * grid_height1
                    cv2.rectangle(f1, (grid_x, grid_y), (grid_x + grid_width1, grid_y + grid_height1), (0, 0, 255), 2)

        for row in range(num_rows):
            for col in range(num_cols):
                if result_matrix2[row, col] == 0:
                    grid_x = roi2_x + col * grid_width2
                    grid_y = roi2_y + row * grid_height2
                    cv2.rectangle(f1, (grid_x, grid_y), (grid_x + grid_width2, grid_y + grid_height2), (0, 0, 255), 2)

        for row in range(num_rows):
            for col in range(num_cols):
                if result_matrix1[row, col] == 1:
                    grid_x = roi1_x + col * grid_width1
                    grid_y = roi1_y + row * grid_height1
                    cv2.rectangle(f1, (grid_x, grid_y), (grid_x + grid_width1, grid_y + grid_height1), (0, 255, 0), 2)

        for row in range(num_rows):
            for col in range(num_cols):
                if result_matrix2[row, col] == 1:
                    grid_x = roi2_x + col * grid_width2
                    grid_y = roi2_y + row * grid_height2
                    cv2.rectangle(f1, (grid_x, grid_y), (grid_x + grid_width2, grid_y + grid_height2), (0, 255, 0), 2)

        density_lane1 = (np.sum(result_matrix1 == 1)) / (num_rows * num_cols)
        density_lane2 = (np.sum(result_matrix2 == 1)) / (num_rows * num_cols)
        density_values_lane1.append(density_lane1)
        density_values_lane2.append(density_lane2)

        if frame_count % 100 == 0:  # Print every 100 frames to reduce console output
            print("Result matrix for frame", frame_count)
            print(result_matrix1)
            print("Density Lane 1:", density_lane1)
            print("Result matrix for lane 2")
            print(result_matrix2)
            print("Density Lane 2:", density_lane2)

        cv2.putText(f1, "Frame: {}".format(frame_count), (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)
        out.write(f1)

        # allow abort if user presses ESC while script has a window focus
        if cv2.waitKey(1) == 27:
            break

    end_time = time.time()
    execution_time = end_time - start_time
    print("Execution Time: {:.2f} seconds".format(execution_time))

    memory_usage = psutil.Process().memory_info().rss
    print("Memory Usage: {:.2f} MB".format(memory_usage / (1024 * 1024)))
    
    avg_density_lane1 = np.mean(density_values_lane1)
    avg_density_lane2 = np.mean(density_values_lane2)
    print("Average Density Lane 1: {:.4f}".format(avg_density_lane1))
    print("Average Density Lane 2: {:.4f}".format(avg_density_lane2))

    if cap is not None:
        try:
            cap.release()
        except Exception:
            pass
    out.release()
    cv2.destroyAllWindows()
    print("frames: "f"{frame_count}")

if __name__ == '__main__':
    profiler = cProfile.Profile()
    profiler.enable()
    main()
    profiler.disable()

    stats = pstats.Stats(profiler).sort_stats('cumtime')
    stats.print_stats(10)  # Print the top 10 functions by cumulative time